"""Parse MFCentral CAS Excel file and import to database."""
import logging
import openpyxl
from mftool import Mftool
from db.database import get_connection

logger = logging.getLogger(__name__)

# Cache for AMFI scheme codes
_amfi_codes_cache = None


def _get_amfi_codes():
    """Get all AMFI scheme codes (cached)."""
    global _amfi_codes_cache
    if _amfi_codes_cache is None:
        mf = Mftool()
        codes = mf.get_scheme_codes()
        # codes is a dict {code: name}
        _amfi_codes_cache = codes if isinstance(codes, dict) else {}
    return _amfi_codes_cache


def _find_amfi_code(scheme_name: str) -> str:
    """Find AMFI code for a scheme by fuzzy name matching."""
    codes = _get_amfi_codes()
    name_lower = scheme_name.lower().strip()

    # Try exact match first
    for code, name in codes.items():
        if name.lower().strip() == name_lower:
            return str(code)

    # Try matching with "direct" and "growth" keywords
    # Extract key parts of scheme name for matching
    keywords = []
    for word in name_lower.replace("-", " ").split():
        if word not in ("fund", "plan", "option", "the", "of", "and", "&", "a"):
            keywords.append(word)

    best_match = None
    best_score = 0
    for code, name in codes.items():
        code_name_lower = name.lower()
        # Count matching keywords
        score = sum(1 for kw in keywords if kw in code_name_lower)
        # Bonus for "direct" and "growth" presence
        if "direct" in code_name_lower and "direct" in name_lower:
            score += 2
        if "growth" in code_name_lower and "growth" in name_lower:
            score += 2
        if score > best_score and score >= len(keywords) * 0.6:
            best_score = score
            best_match = str(code)

    return best_match or ""


def parse_cas_excel(excel_path: str) -> dict:
    """Parse MFCentral CAS Excel file (Portfolio Details + Transaction Details).

    Returns dict with keys: investor, holdings, transactions
    """
    logger.info(f"Parsing CAS Excel: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    result = {"investor": {}, "holdings": [], "transactions": []}

    # --- Portfolio Details sheet ---
    if "Portfolio Details" in wb.sheetnames:
        ws = wb["Portfolio Details"]
        rows = list(ws.iter_rows(values_only=True))

        # Parse investor info (rows 1-4)
        for row in rows[:5]:
            if row and row[0]:
                key = str(row[0]).strip()
                val = str(row[1]).strip() if row[1] else ""
                if key == "Name":
                    result["investor"]["name"] = val
                elif key == "PAN":
                    result["investor"]["pan"] = val
                elif key == "Email":
                    result["investor"]["email"] = val

        # Find header row for holdings
        header_idx = None
        for i, row in enumerate(rows):
            if row and row[0] and str(row[0]).strip() == "Scheme Name":
                header_idx = i
                break

        if header_idx is not None:
            # Headers: Scheme Name, AMC Name, Category, Folio No., Invested Value, Current Value, Returns, Units
            for row in rows[header_idx + 1:]:
                if not row or not row[0] or str(row[0]).strip() == "":
                    continue
                scheme_name = str(row[0]).strip() if row[0] else ""
                amc = str(row[1]).strip() if row[1] else ""
                category = str(row[2]).strip() if row[2] else ""
                folio = str(row[3]).strip() if row[3] else ""
                invested = float(row[4]) if row[4] else 0
                current_val = float(row[5]) if row[5] else 0
                returns = float(row[6]) if row[6] else 0
                units = float(row[7]) if row[7] else 0

                if units == 0 and current_val == 0:
                    continue  # Skip zero-balance schemes

                nav = current_val / units if units > 0 else 0

                result["holdings"].append({
                    "scheme_name": scheme_name,
                    "amc": amc,
                    "category": category,
                    "folio": folio,
                    "invested_value": invested,
                    "current_value": current_val,
                    "returns": returns,
                    "units": units,
                    "nav": nav,
                })

    # --- Transaction Details sheet ---
    if "Transaction Details" in wb.sheetnames:
        ws = wb["Transaction Details"]
        rows = list(ws.iter_rows(values_only=True))

        # Find header row
        header_idx = None
        for i, row in enumerate(rows):
            if row and row[0] and str(row[0]).strip() == "Scheme Name":
                header_idx = i
                break

        if header_idx is not None:
            # Headers: Scheme Name, Transaction Description, Date, NAV, Units, Amount
            for row in rows[header_idx + 1:]:
                if not row or not row[0] or str(row[0]).strip() == "":
                    continue
                scheme_name = str(row[0]).strip() if row[0] else ""
                description = str(row[1]).strip() if row[1] else ""
                date_val = str(row[2]).strip() if row[2] else ""
                nav = float(row[3]) if row[3] else 0
                units = float(row[4]) if row[4] else 0
                amount = float(row[5]) if row[5] else 0

                # Skip non-financial transactions
                if nav == 0 and units == 0 and amount == 0:
                    continue

                result["transactions"].append({
                    "scheme_name": scheme_name,
                    "description": description,
                    "date": date_val,
                    "nav": nav,
                    "units": units,
                    "amount": amount,
                })

    wb.close()
    logger.info(f"Parsed {len(result['holdings'])} holdings, {len(result['transactions'])} transactions")
    return result


def import_cas_excel_to_db(excel_path: str):
    """Parse CAS Excel and import holdings + transactions into database."""
    data = parse_cas_excel(excel_path)
    conn = get_connection()
    cursor = conn.cursor()

    print("Looking up AMFI codes for schemes...")
    imported_schemes = 0
    imported_tx = 0
    amfi_found = 0

    for h in data["holdings"]:
        folio = h["folio"]
        scheme_name = h["scheme_name"]
        amc = h["amc"]
        units = h["units"]
        invested = h["invested_value"]
        current_val = h["current_value"]
        nav = h["nav"]
        category = h["category"]
        is_active = 1 if units > 0 else 0

        # Look up AMFI code
        amfi_code = _find_amfi_code(scheme_name)
        if amfi_code:
            amfi_found += 1

        cursor.execute("""
            INSERT INTO holdings (folio, amc, scheme_name, scheme_type, amfi_code,
                current_units, latest_nav, latest_value, cost_value, is_active, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(folio, scheme_name) DO UPDATE SET
                amc = excluded.amc,
                scheme_type = excluded.scheme_type,
                amfi_code = CASE WHEN excluded.amfi_code != '' THEN excluded.amfi_code
                                 ELSE holdings.amfi_code END,
                current_units = excluded.current_units,
                latest_nav = excluded.latest_nav,
                latest_value = excluded.latest_value,
                cost_value = excluded.cost_value,
                is_active = excluded.is_active,
                updated_at = CURRENT_TIMESTAMP
        """, (folio, amc, scheme_name, category, amfi_code, units, nav,
              current_val, invested, is_active))
        imported_schemes += 1

    # Import transactions
    for tx in data["transactions"]:
        # Find matching holding
        cursor.execute(
            "SELECT id, folio FROM holdings WHERE scheme_name LIKE ?",
            (f"%{tx['scheme_name'][:40]}%",)
        )
        match = cursor.fetchone()
        holding_id = match["id"] if match else None
        folio = match["folio"] if match else ""

        # Determine tx_type from description
        desc_lower = tx["description"].lower()
        if "purchase" in desc_lower or "sip" in desc_lower:
            tx_type = "PURCHASE"
        elif "redemption" in desc_lower or "withdraw" in desc_lower:
            tx_type = "REDEMPTION"
        elif "switch" in desc_lower and "out" in desc_lower:
            tx_type = "SWITCH_OUT"
        elif "switch" in desc_lower and "in" in desc_lower:
            tx_type = "SWITCH_IN"
        else:
            tx_type = "MISC"

        # Normalize date (DD-MMM-YYYY → YYYY-MM-DD)
        date_str = tx["date"]
        try:
            from datetime import datetime
            for fmt in ("%d-%b-%Y", "%d-%m-%Y", "%Y-%m-%d"):
                try:
                    date_str = datetime.strptime(date_str, fmt).strftime("%Y-%m-%d")
                    break
                except ValueError:
                    continue
        except Exception:
            pass

        try:
            cursor.execute("""
                INSERT INTO transactions (holding_id, folio, scheme_name, date,
                    description, amount, units, nav, tx_type)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(folio, scheme_name, date, amount, units, tx_type) DO NOTHING
            """, (holding_id, folio, tx["scheme_name"], date_str,
                  tx["description"], tx["amount"], tx["units"], tx["nav"], tx_type))
            if cursor.rowcount > 0:
                imported_tx += 1
        except Exception as e:
            logger.warning(f"Failed to insert transaction: {e}")

    conn.commit()
    conn.close()

    print(f"\n{'='*50}")
    print(f"CAS Excel Import Summary")
    print(f"{'='*50}")
    print(f"Schemes imported : {imported_schemes}")
    print(f"AMFI codes found : {amfi_found}/{imported_schemes}")
    print(f"Transactions     : {imported_tx}")
    print(f"{'='*50}\n")

    return {"schemes": imported_schemes, "transactions": imported_tx}
