"""Generate daily portfolio Excel report with Stocks and MF sheets."""
import os
import sys
from datetime import date, datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from db.database import get_connection, init_db

REPORT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "portfolio_report.xlsx")

# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
DATE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
GREEN_FONT = Font(color="006100")
RED_FONT = Font(color="9C0006")
NUM_FMT = '#,##0.00'
PCT_FMT = '0.00%'


def _style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 30)


def generate_stocks_sheet(wb):
    """Create/update Stocks sheet with daily price history."""
    conn = get_connection()
    cur = conn.cursor()

    # Get active stock holdings
    cur.execute("""
        SELECT ticker, security_name, isin, quantity, invested_value,
               latest_price, latest_value
        FROM stock_holdings
        WHERE is_active = 1 AND ticker IS NOT NULL AND ticker != ''
        ORDER BY security_name
    """)
    holdings = [dict(r) for r in cur.fetchall()]

    if not holdings:
        conn.close()
        return

    tickers = [h["ticker"] for h in holdings]

    # Get all price dates
    placeholders = ",".join("?" * len(tickers))
    cur.execute(f"""
        SELECT DISTINCT date FROM stock_prices
        WHERE ticker IN ({placeholders})
        ORDER BY date
    """, tickers)
    dates = [r["date"] for r in cur.fetchall()]

    # Build price lookup: {ticker: {date: price}}
    cur.execute(f"""
        SELECT ticker, date, close_price FROM stock_prices
        WHERE ticker IN ({placeholders})
        ORDER BY ticker, date
    """, tickers)
    prices = {}
    for r in cur.fetchall():
        prices.setdefault(r["ticker"], {})[r["date"]] = r["close_price"]

    conn.close()

    # Create or replace sheet
    if "Stocks" in wb.sheetnames:
        del wb["Stocks"]
    ws = wb.create_sheet("Stocks", 0)

    # Headers: Stock | Ticker | Qty | Invested | Latest Price | Latest Value | P&L | date1 | date2 ...
    static_cols = ["Stock", "Ticker", "Qty", "Invested (₹)", "Latest Price (₹)",
                   "Latest Value (₹)", "P&L (₹)"]
    headers = static_cols + dates

    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)

    _style_header(ws, 1, len(headers))

    # Style date columns
    for col_idx in range(len(static_cols) + 1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = DATE_FILL
        cell.font = Font(bold=True, size=10)

    # Data rows
    for row_idx, h in enumerate(holdings, 2):
        ws.cell(row=row_idx, column=1, value=h["security_name"])
        ws.cell(row=row_idx, column=2, value=h["ticker"].replace(".NS", ""))
        ws.cell(row=row_idx, column=3, value=h["quantity"])
        ws.cell(row=row_idx, column=4, value=h["invested_value"] or 0).number_format = NUM_FMT
        ws.cell(row=row_idx, column=5, value=h["latest_price"] or 0).number_format = NUM_FMT
        ws.cell(row=row_idx, column=6, value=h["latest_value"] or 0).number_format = NUM_FMT

        pnl = (h["latest_value"] or 0) - (h["invested_value"] or 0)
        pnl_cell = ws.cell(row=row_idx, column=7, value=pnl)
        pnl_cell.number_format = NUM_FMT
        pnl_cell.font = GREEN_FONT if pnl >= 0 else RED_FONT

        # Daily prices
        ticker_prices = prices.get(h["ticker"], {})
        for d_idx, d in enumerate(dates):
            col = len(static_cols) + 1 + d_idx
            price = ticker_prices.get(d)
            if price is not None:
                ws.cell(row=row_idx, column=col, value=price).number_format = NUM_FMT

    # Totals row
    total_row = len(holdings) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=4,
            value=sum(h["invested_value"] or 0 for h in holdings)).number_format = NUM_FMT
    ws.cell(row=total_row, column=6,
            value=sum(h["latest_value"] or 0 for h in holdings)).number_format = NUM_FMT
    total_pnl = sum((h["latest_value"] or 0) - (h["invested_value"] or 0) for h in holdings)
    t_cell = ws.cell(row=total_row, column=7, value=total_pnl)
    t_cell.number_format = NUM_FMT
    t_cell.font = Font(bold=True, color="006100" if total_pnl >= 0 else "9C0006")

    _auto_width(ws)
    ws.freeze_panes = "C2"


def generate_mf_sheet(wb):
    """Create/update Mutual Funds sheet with daily NAV history."""
    conn = get_connection()
    cur = conn.cursor()

    # Get active MF holdings
    cur.execute("""
        SELECT amfi_code, scheme_name, folio, current_units, cost_value,
               latest_nav, latest_value
        FROM holdings
        WHERE is_active = 1 AND amfi_code IS NOT NULL AND amfi_code != ''
        ORDER BY scheme_name
    """)
    holdings = [dict(r) for r in cur.fetchall()]

    if not holdings:
        conn.close()
        return

    amfi_codes = [h["amfi_code"] for h in holdings]

    # Get all NAV dates
    placeholders = ",".join("?" * len(amfi_codes))
    cur.execute(f"""
        SELECT DISTINCT date FROM daily_prices
        WHERE amfi_code IN ({placeholders})
        ORDER BY date
    """, amfi_codes)
    dates = [r["date"] for r in cur.fetchall()]

    # Build NAV lookup: {amfi_code: {date: nav}}
    cur.execute(f"""
        SELECT amfi_code, date, nav FROM daily_prices
        WHERE amfi_code IN ({placeholders})
        ORDER BY amfi_code, date
    """, amfi_codes)
    navs = {}
    for r in cur.fetchall():
        navs.setdefault(r["amfi_code"], {})[r["date"]] = r["nav"]

    conn.close()

    # Create or replace sheet
    if "Mutual Funds" in wb.sheetnames:
        del wb["Mutual Funds"]
    ws = wb.create_sheet("Mutual Funds", 1)

    # Headers
    static_cols = ["Scheme", "AMFI Code", "Folio", "Units", "Cost (₹)",
                   "Latest NAV", "Current Value (₹)", "P&L (₹)", "Returns %"]
    headers = static_cols + dates

    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)

    _style_header(ws, 1, len(headers))

    for col_idx in range(len(static_cols) + 1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = DATE_FILL
        cell.font = Font(bold=True, size=10)

    # Data rows
    for row_idx, h in enumerate(holdings, 2):
        ws.cell(row=row_idx, column=1, value=h["scheme_name"])
        ws.cell(row=row_idx, column=2, value=h["amfi_code"])
        ws.cell(row=row_idx, column=3, value=h["folio"])
        ws.cell(row=row_idx, column=4, value=h["current_units"]).number_format = '0.000'
        ws.cell(row=row_idx, column=5, value=h["cost_value"] or 0).number_format = NUM_FMT
        ws.cell(row=row_idx, column=6, value=h["latest_nav"] or 0).number_format = NUM_FMT
        ws.cell(row=row_idx, column=7, value=h["latest_value"] or 0).number_format = NUM_FMT

        pnl = (h["latest_value"] or 0) - (h["cost_value"] or 0)
        pnl_cell = ws.cell(row=row_idx, column=8, value=pnl)
        pnl_cell.number_format = NUM_FMT
        pnl_cell.font = GREEN_FONT if pnl >= 0 else RED_FONT

        ret_pct = pnl / h["cost_value"] if h["cost_value"] else 0
        ret_cell = ws.cell(row=row_idx, column=9, value=ret_pct)
        ret_cell.number_format = PCT_FMT
        ret_cell.font = GREEN_FONT if ret_pct >= 0 else RED_FONT

        # Daily NAVs
        code_navs = navs.get(h["amfi_code"], {})
        for d_idx, d in enumerate(dates):
            col = len(static_cols) + 1 + d_idx
            nav = code_navs.get(d)
            if nav is not None:
                ws.cell(row=row_idx, column=col, value=nav).number_format = NUM_FMT

    # Totals row
    total_row = len(holdings) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=5,
            value=sum(h["cost_value"] or 0 for h in holdings)).number_format = NUM_FMT
    ws.cell(row=total_row, column=7,
            value=sum(h["latest_value"] or 0 for h in holdings)).number_format = NUM_FMT
    total_pnl = sum((h["latest_value"] or 0) - (h["cost_value"] or 0) for h in holdings)
    t_cell = ws.cell(row=total_row, column=8, value=total_pnl)
    t_cell.number_format = NUM_FMT
    t_cell.font = Font(bold=True, color="006100" if total_pnl >= 0 else "9C0006")
    total_cost = sum(h["cost_value"] or 0 for h in holdings)
    ret = total_pnl / total_cost if total_cost else 0
    ws.cell(row=total_row, column=9, value=ret).number_format = PCT_FMT

    _auto_width(ws)
    ws.freeze_panes = "B2"


def generate_report():
    """Generate or update the portfolio Excel report."""
    init_db()

    wb = openpyxl.Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    generate_stocks_sheet(wb)
    generate_mf_sheet(wb)

    # Add a Summary sheet
    ws = wb.create_sheet("Summary", 0)
    conn = get_connection()
    cur = conn.cursor()

    ws.cell(row=1, column=1, value="Portfolio Summary").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M IST')}")

    # MF totals
    cur.execute("SELECT COUNT(*) as cnt, COALESCE(SUM(cost_value),0) as cost, COALESCE(SUM(latest_value),0) as val FROM holdings WHERE is_active=1")
    mf = cur.fetchone()

    # Stock totals
    cur.execute("SELECT COUNT(*) as cnt, COALESCE(SUM(invested_value),0) as cost, COALESCE(SUM(latest_value),0) as val FROM stock_holdings WHERE is_active=1")
    stk = cur.fetchone()

    row = 4
    labels = [
        ("", "Count", "Invested (₹)", "Current Value (₹)", "P&L (₹)", "Returns"),
        ("Mutual Funds", mf["cnt"], mf["cost"], mf["val"],
         mf["val"] - mf["cost"],
         (mf["val"] - mf["cost"]) / mf["cost"] if mf["cost"] else 0),
        ("Stocks", stk["cnt"], stk["cost"], stk["val"],
         stk["val"] - stk["cost"],
         (stk["val"] - stk["cost"]) / stk["cost"] if stk["cost"] else 0),
        ("TOTAL", mf["cnt"] + stk["cnt"],
         mf["cost"] + stk["cost"], mf["val"] + stk["val"],
         (mf["val"] + stk["val"]) - (mf["cost"] + stk["cost"]),
         ((mf["val"] + stk["val"]) - (mf["cost"] + stk["cost"])) / (mf["cost"] + stk["cost"])
         if (mf["cost"] + stk["cost"]) else 0),
    ]

    for i, vals in enumerate(labels):
        for j, v in enumerate(vals):
            cell = ws.cell(row=row + i, column=j + 1, value=v)
            if i == 0:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
            elif j >= 2 and j <= 4:
                cell.number_format = NUM_FMT
            elif j == 5 and i > 0:
                cell.number_format = PCT_FMT
            if i > 0 and j == 4:
                cell.font = GREEN_FONT if v >= 0 else RED_FONT

    # Last update info
    cur.execute("SELECT date FROM portfolio_snapshots ORDER BY date DESC LIMIT 1")
    snap = cur.fetchone()
    if snap:
        ws.cell(row=row + len(labels) + 1, column=1,
                value=f"Last snapshot: {snap['date']}")

    conn.close()
    _auto_width(ws)

    wb.save(REPORT_PATH)
    print(f"Portfolio report saved: {REPORT_PATH}")
    return REPORT_PATH


if __name__ == "__main__":
    path = generate_report()
    print(f"Done: {path}")
